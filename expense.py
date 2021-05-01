import os
import sys
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from matplotlib import pyplot as plt
from matplotlib import ticker

pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.width', 320)
pd.set_option('display.max_columns', 30)
pd.set_option('display.max_rows', 500)

file_name = 'expense.xlsx'
menu_ls = ['Insert Record', 'Delete Record', 'Create Sheet', 'Delete Sheet', 'View Sheet', 'View Chart', 'Exit']
sheet_header = ['Date', 'Category', 'Amount', 'Note']
category_ls = ['grocery/food', 'utility', 'monthly fee', 'rent', 'other', 'income']
chart_ls = ['Expense Summary - Pie', 'Expense Summary - Bar', 'Balance Summary', 'All']

if not os.path.exists(file_name):
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    writer.save()


def get_platform():
    platforms = {
        'linux1': 'Linux',
        'linux2': 'Linux',
        'darwin': 'OS X',
        'win32': 'Windows'
    }
    if sys.platform not in platforms:
        return sys.platform

    return platforms[sys.platform]


def clear():
    platform_name = get_platform()
    if platform_name == 'Windows':
        os.system('cls')
    else:
        os.system('clear')


def insert_record(sheet_name):
    if sheet_name != 'All Sheets':
        print('Original Sheet')
        display_sheet(sheet_name)
        book = load_workbook(file_name)
        sheet = book[sheet_name]
        # activate sheet
        for s in range(len(book.sheetnames)):
            if book.sheetnames[s] == sheet_name:
                book.active = s
                break
        print('\nEnter record')
        date = enter_date()
        category = select_category()
        amount = enter_amount()
        if category != 'income':
            amount = -amount
        note = input_esc('Note: ')
        record = [date, category, amount, note]
        sheet.append(record)
        book.save(file_name)
    print('\nUpdated Sheet')
    display_sheet(sheet_name)
    insert_record(continue_edit())


def delete_record(sheet_name):
    if sheet_name != 'All Sheets':
        print('Original Sheet')
        row_num_dict = display_sheet(sheet_name)
        book = load_workbook(file_name)
        sheet = book[sheet_name]
        # activate sheet
        for s in range(len(book.sheetnames)):
            if book.sheetnames[s] == sheet_name:
                book.active = s
                break
        sheet_size = sheet.max_row-1
        ans_row_del = input_esc('\nSelect row number to delete: ')
        if ans_row_del in int_str_list(sheet_size):
            row_num_del = row_num_dict.get(int(ans_row_del))
            sheet.delete_rows(row_num_del, 1)
            book.save(file_name)
        else:
            clear()
            print('\nInvalid input, please re-enter\n')
            delete_record(sheet_name)
    print('\nUpdated Sheet')
    display_sheet(sheet_name)
    delete_record(continue_edit())


def create_sheet():
    # sheet list before creating a new sheet
    print('Original Sheet List')
    view_sheet_list()

    book = load_workbook(file_name)
    ans_sheet_name = input_esc('New Sheet Name: ')
    book.create_sheet(ans_sheet_name)
    book.save(file_name)

    book = load_workbook(file_name)
    sheet = book[ans_sheet_name]
    sheet.append(sheet_header)
    book.save(file_name)

    book = load_workbook(file_name)
    if 'Sheet1' in book.sheetnames:
        del book['Sheet1']
    book.save(file_name)

    # sheet list after creating a new sheet
    print('\nUpdated Sheet List')
    view_sheet_list()

    cont_process = continue_process()
    if cont_process is True:
        create_sheet()
    else:
        menu()


def delete_sheet(sheet_name):
    if sheet_name != 'All Sheets':
        book = load_workbook(file_name)
        del book[sheet_name]
        book.save(file_name)
        print(sheet_name + ' has been deleted...')

    # sheet list after deleting a new sheet
    print('\nUpdated Sheet List')
    view_sheet_list()
    delete_sheet(continue_view())


def view_sheet(sheet_name):
    if sheet_name != 'All Sheets':
        sheet_view = pd.read_excel(file_name, sheet_name)
    else:
        book = load_workbook(file_name)
        sheet_ls = book.sheetnames
        sheet_view = pd.DataFrame()
        for name in sheet_ls:
            sheet = pd.read_excel(file_name, name)
            sheet_view = sheet_view.append(sheet, ignore_index=True)
    sheet_view.sort_values(by=['Date', 'Category', 'Amount', 'Note'], ascending=True, inplace=True, ignore_index=True)
    sheet_view.index += 1
    print(sheet_view)
    # view balance summary
    if sheet_name != 'All Sheets':
        print('\n' + sheet_name + ' Balance Summary')
    else:
        print('\nMonthly Balance Summary')
    sheet_view['Amount'] = abs(sheet_view['Amount'])
    sheet_view['Date'] = sheet_view['Date'].dt.strftime('%Y-%m')
    income_summary = sheet_view[sheet_view['Category'] == 'income']
    income_summary.rename(columns={'Amount': 'Income'}, inplace=True)
    total_income = income_summary.groupby(['Date'], as_index=False).agg({'Income': 'sum'})
    expense_summary = sheet_view[sheet_view['Category'] != 'income']
    expense_summary.rename(columns={'Amount': 'Expense'}, inplace=True)
    expense_summary['Category'] = 'Expense'
    total_expense = expense_summary.groupby(['Date'], as_index=False).agg({'Expense': 'sum'})
    balance_summary = pd.merge(
        total_income,
        total_expense,
        how='outer',
        on='Date'
    )
    balance_summary['Balance'] = balance_summary['Income'] - balance_summary['Expense']
    if sheet_name == 'All Sheets':
        balance_summary = balance_summary.append(balance_summary.sum(numeric_only=True), ignore_index=True)
        balance_summary.fillna('TOTAL', inplace=True)
    print(balance_summary)
    view_sheet(continue_view())


def view_charts(sheet_name, chart_name):
    display_sheet(sheet_name)
    if chart_name == 'Expense Summary - Pie':
        pie_chart(sheet_name)
    elif chart_name == 'Expense Summary - Bar':
        stacked_bar_chart(sheet_name)
    elif chart_name == 'Balance Summary':
        balance_bar_chart(sheet_name)
    else:
        pie_chart(sheet_name)
        stacked_bar_chart(sheet_name)
        balance_bar_chart(sheet_name)
    plt.show()
    input('Press Enter to continue...')
    clear()
    view_charts(continue_view(), select_chart())


# select worksheet to edit
def select_sheet():
    book = load_workbook(file_name)
    sheet_ls = book.sheetnames
    sheet_ls.insert(0, 'Go Back to Main Menu')
    sheet_ls.append('All Sheets')
    for i in range(len(sheet_ls)):
        print(str(i) + ': ' + sheet_ls[i])
        i += 1
    ans_sheet = input_esc('\nPlease select a worksheet: ')
    clear()
    if ans_sheet in int_str_list(len(sheet_ls)-1):
        sheet_name = sheet_ls[int(ans_sheet)]
        return sheet_name
    elif ans_sheet == '0':
        menu()
    else:
        print('Invalid input, please re-select\n')
        sheet_name = select_sheet()
        return sheet_name


def view_sheet_list():
    book = load_workbook(file_name)
    sheet_ls = book.sheetnames
    for i in range(len(sheet_ls)):
        print(str(i+1) + ': ' + sheet_ls[i])
        i += 1


def display_sheet(sheet_name):
    if sheet_name != 'All Sheets':
        sheet_view = pd.read_excel(file_name, sheet_name)
    else:
        book = load_workbook(file_name)
        sheet_ls = book.sheetnames
        sheet_view = pd.DataFrame()
        for name in sheet_ls:
            sheet = pd.read_excel(file_name, name)
            sheet_view = sheet_view.append(sheet, ignore_index=True)
    sheet_view['RowNum'] = sheet_view.index + 2
    sheet_view.sort_values(by=['Date', 'Category', 'Amount', 'Note'], ascending=True, inplace=True, ignore_index=True)
    sheet_view.index += 1
    row_num_dict = pd.Series(sheet_view.RowNum.values, index=sheet_view.index).to_dict()
    sheet_view.drop(['RowNum'], axis=1, inplace=True)
    print(sheet_view)
    return row_num_dict


def enter_date():
    ans_date = input_esc('Date (YYYYmmdd): ')
    if date_error(ans_date):
        date = enter_date()
    else:
        date = datetime.strptime(ans_date, '%Y%m%d').date()
        print('\n')
    return date


def enter_amount():
    ans_amount = input_esc('Amount: ')
    if float_error(ans_amount):
        amount = enter_amount()
    else:
        amount = float(ans_amount)
        print('\n')
    return amount


def select_category():
    print('Category List')
    for i in range(len(category_ls)):
        print(str(i+1) + ': ' + category_ls[i])
        i += 1
    ans_category = input_esc('Category: ')
    if ans_category in int_str_list(len(category_ls)):
        category = category_ls[int(ans_category)-1]
        print('\n')
    else:
        print('\nInvalid input, please re-select\n')
        category = select_category()
    return category


def select_chart():
    print('Chart List')
    for i in range(len(chart_ls)):
        print(str(i+1) + ': ' + chart_ls[i])
        i += 1
    ans_chart = input_esc('Please select: ')
    if ans_chart in int_str_list(len(chart_ls)):
        chart_name = chart_ls[int(ans_chart)-1]
        clear()
        return chart_name
    else:
        print('\nInvalid input, please re-select\n')
        chart_name = select_chart()
        return chart_name


def continue_edit():
    ans_cont_edit = input('\nPress 1 to edit another record; press 2 to switch a sheet; press 3 to exit: ')
    if ans_cont_edit == '1':
        clear()
        book = load_workbook(file_name)
        sheet_name = book.active.title
        return sheet_name
    elif ans_cont_edit == '2':
        clear()
        sheet_name = select_sheet()
        return sheet_name
    elif ans_cont_edit == '3':
        clear()
        menu()
    else:
        sheet_name = continue_edit()
        return sheet_name


def continue_view():
    ans_cont_view = input('\nPress 1 to select another sheet; press 2 to exit: ')
    if ans_cont_view == '1':
        clear()
        sheet_name = select_sheet()
        return sheet_name
    elif ans_cont_view == '2':
        clear()
        menu()
    else:
        sheet_name = continue_view()
        return sheet_name


def continue_process():
    ans_cont_process = input('\nPress 1 to continue; press 2 to exit: ')
    if ans_cont_process == '1':
        clear()
        cont_process = True
    elif ans_cont_process == '2':
        clear()
        cont_process = False
    else:
        cont_process = continue_process()
    return cont_process


def data_prep_sgl(sheet_name):
    if sheet_name != 'All Sheets':
        expense_df = pd.read_excel(file_name, sheet_name)[['Date', 'Category', 'Amount']]
    else:
        book = load_workbook(file_name)
        sheet_ls = book.sheetnames
        expense_df = pd.DataFrame()
        for name in sheet_ls:
            sheet = pd.read_excel(file_name, name)
            expense_df = expense_df.append(sheet, ignore_index=True)
        expense_df['Date'] = expense_df['Date'].dt.strftime('%Y-%m')
    expense_df['Amount'] = abs(expense_df['Amount'])
    expense_df = expense_df.groupby(['Date', 'Category'])['Amount'].sum().reset_index()
    expense_df = expense_df.pivot_table(index='Date', columns='Category', values=['Amount'], fill_value=0)
    expense_df.reset_index(inplace=True)
    daily_exp_columns = [s2 for (s1, s2) in expense_df.columns.tolist()]
    daily_exp_columns[daily_exp_columns.index('')] = 'date'
    expense_df.columns = daily_exp_columns
    columns_to_add = [i for i in category_ls if i not in expense_df.columns]
    if columns_to_add:
        for col in columns_to_add:
            expense_df[col] = 0
    expense_df = expense_df[[c for c in expense_df if c not in ['other', 'income']] + ['other', 'income']]
    return expense_df


def pie_chart(sheet_name):
    expense_df = data_prep_sgl(sheet_name)
    expense_df.drop(['date', 'income'], axis=1, inplace=True)
    sum_expense = expense_df.sum(axis=0).to_frame(name='')
    sum_expense.reset_index(level=0, inplace=True)
    sum_expense.rename(columns={'index': 'Category'}, inplace=True)
    sum_expense.groupby(['Category']).sum().plot(kind='pie', subplots=True, shadow=False, startangle=0, figsize=(14, 8), autopct='%1.1f%%')
    if sheet_name != 'All Sheets':
        chart_title = sheet_name + ' Expense Summary by Category'
    else:
        chart_title = 'Expense Summary by Category'
    plt.title(chart_title, fontdict={'fontweight': 'bold', 'fontsize': 18})
    plt.legend(bbox_to_anchor=(1, 1), borderaxespad=0., fontsize=10, fancybox=True, shadow=False)
    # plt.show()


def stacked_bar_chart(sheet_name):
    expense_df = data_prep_sgl(sheet_name)
    expense_df.drop(['rent', 'income'], axis=1, inplace=True)
    if sheet_name != 'All Sheets':
        expense_df = expense_df.set_index('date').asfreq('1D', fill_value=0)
    else:
        expense_df = expense_df.set_index('date')
    plt.style.use('ggplot')
    ax = expense_df.plot(stacked=True, kind='bar', figsize=(14, 8), rot='horizontal')
    # .patches is everything inside of the chart
    for rect in ax.patches:
        # Find where everything is located
        height = rect.get_height()
        width = rect.get_width()
        x = rect.get_x()
        y = rect.get_y()
        # The height of the bar is the data value and can be used as the label
        label_text = f'{height:.2f}'
        label_x = x + width / 2
        label_y = y + height / 2
        # plot only when height is greater than specified value
        if height > 0:
            ax.text(label_x, label_y, label_text, ha='center', va='center', fontsize=7)
    plt.legend(bbox_to_anchor=(1, 1), borderaxespad=0., fontsize=10, fancybox=True, shadow=False)
    ax.set_ylabel("Amount (C$)", fontsize=10)
    ax.set_xlabel("Date", fontsize=10)
    plt.xticks(rotation=30)
    if sheet_name != 'All Sheets':
        chart_title = sheet_name + ' Daily Expense Summary'
    else:
        chart_title = 'Monthly Expense Summary'
    plt.title(chart_title, fontdict={'fontweight': 'bold', 'fontsize': 18})
    tick_labels = expense_df.index.tolist()
    if sheet_name != 'All Sheets':
        for (i, item) in enumerate(tick_labels):
            if i == 0:
                tick_labels[i] = item.strftime('%Y %b %d')
            elif item.strftime('%d') == '01':
                tick_labels[i] = item.strftime('%b %d')
            elif item.strftime('%b %d') == 'Jan 01':
                tick_labels[i] = item.strftime('%Y %b %d')
            else:
                tick_labels[i] = item.strftime('%d')
    else:
        for (i, item) in enumerate(tick_labels):
            if i == 0:
                tick_labels[i] = item
            elif item[-2:] == '01':
                tick_labels[i] = item
            else:
                tick_labels[i] = item[-2:]
    ax.xaxis.set_major_formatter(ticker.FixedFormatter(tick_labels))
    plt.gcf().autofmt_xdate()
    # plt.show()


def balance_bar_chart(sheet_name):
    expense_df = data_prep_sgl(sheet_name)
    expense_df = expense_df.melt(id_vars=['date'], var_name='Category', value_name='Amount')
    expense_df = expense_df[expense_df['Amount'] != 0]
    if sheet_name != 'All Sheets':
        expense_df['Month'] = expense_df['date'].dt.strftime('%Y-%m')
    else:
        expense_df['Month'] = expense_df['date']
    income_agg = expense_df[expense_df['Category'] == 'income']
    income_agg.rename(columns={'Amount': 'Income'}, inplace=True)
    expense_agg = expense_df[expense_df['Category'] != 'income']
    expense_agg.rename(columns={'Amount': 'Expense'}, inplace=True)
    expense_cat_mon = expense_agg.groupby(['Category', 'Month'], as_index=False).agg({'Expense': 'sum'})
    expense_mon = expense_cat_mon.groupby(['Month'], as_index=False).agg({'Expense': 'sum'})
    income_mon = income_agg.groupby(['Month'], as_index=False).agg({'Income': 'sum'})
    balance_mon = pd.merge(
        income_mon,
        expense_mon,
        how='outer',
        on='Month'
    )
    balance_mon['Balance'] = balance_mon['Income'] - balance_mon['Expense']
    balance_mon.loc[(balance_mon['Month'].str[-2:] != '01') & (balance_mon.index != 0), 'Month'] = balance_mon['Month'].str[-2:]
    ax = balance_mon.plot(x='Month', y=['Income', 'Expense', 'Balance'], kind='bar', figsize=(14, 8))
    # .patches is everything inside of the chart
    for rect in ax.patches:
        # Find where everything is located
        height = rect.get_height()
        width = rect.get_width()
        x = rect.get_x()
        y = rect.get_y()
        # The height of the bar is the data value and can be used as the label
        label_text = f'{height:.2f}'
        label_x = x + width / 2
        label_y = y + height / 2
        # plot only when height is greater than specified value
        if height > -100000000000:
            ax.text(label_x, label_y, label_text, ha='center', va='center', fontsize=7)
    ax.set_ylabel("Amount (C$)", fontsize=10)
    ax.set_xlabel("Month", fontsize=10)
    plt.legend(bbox_to_anchor=(1, 1), borderaxespad=0., fontsize=10, fancybox=True, shadow=False)
    if sheet_name != 'All Sheets':
        chart_title = sheet_name + ' Balance Summary'
    else:
        chart_title = 'Monthly Balance Summary'
    plt.title(chart_title, fontdict={'fontweight': 'bold', 'fontsize': 18})
    plt.xticks(rotation=30)
    # plt.show()


def int_error(msg):
    try:
        int(msg)
        return False
    except ValueError:
        print('\nInvalid input, please re-enter\n')
        return True


def float_error(msg):
    try:
        float(msg)
        return False
    except ValueError:
        print('\nInvalid input, please re-enter\n')
        return True


def date_error(msg):
    try:
        datetime.strptime(msg, '%Y%m%d').date()
        return False
    except ValueError:
        print('\nInvalid input, please re-enter\n')
        return True


def input_esc(msg):
    ans = input(msg)
    if ans == 'esc':
        clear()
        menu()
    else:
        return ans
    
    
def int_str_list(length):
    str_list = list(map(str, list(range(length+1))[1:]))
    return str_list


def menu():
    print('Main Menu')
    for i in range(len(menu_ls)):
        print(str(i+1) + ': ' + menu_ls[i])
        i += 1
    ans_menu = input('\nPlease select: ')
    clear()
    if ans_menu in int_str_list(len(menu_ls)):
        menu_item = menu_ls[int(ans_menu)-1]

        if menu_item == 'Insert Record':
            insert_record(select_sheet())
        elif menu_item == 'Delete Record':
            delete_record(select_sheet())
        elif menu_item == 'Create Sheet':
            create_sheet()
        elif menu_item == 'Delete Sheet':
            delete_sheet(select_sheet())
        elif menu_item == 'View Sheet':
            view_sheet(select_sheet())
        elif menu_item == 'View Chart':
            view_charts(select_sheet(), select_chart())
        elif menu_item == 'Exit':
            quit()

    else:
        clear()
        print('Invalid input, please re-select\n')
        menu()
        

def main_code():
    clear()
    menu()


if __name__ == '__main__':
    main_code()
